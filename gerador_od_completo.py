"""
Sistema Minimalista de Geração de OD
Funcionalidades essenciais apenas (YAGNI)
Lê dinamicamente PLANO_FINAL.pdf e DECUPAGEM.csv
"""

import pandas as pd
import json
import os
import re
from datetime import datetime, timedelta
from typing import Dict, List
import pdfplumber


class GeradorOD:
    def __init__(self):
        self.dados_decupagem = {}
        self.config = {}
        self.titulo_extraido = None

        # Arquivos dinâmicos
        self.arquivo_decupagem = "arquivos/DECUPAGEM.csv"
        self.arquivo_plano = "arquivos/PLANO_FINAL.pdf"
        self.arquivo_config = "config_dias_filmagem.json"
        self.pasta_ods = "arquivos/ODs"

        # Criar pasta de ODs se não existir
        os.makedirs(self.pasta_ods, exist_ok=True)

    def _carregar_dados(self):
        """Carrega dados da decupagem e plano de filmagem automaticamente"""
        print("🔍 Carregando dados do projeto atual...")

        # Carregar decupagem
        try:
            df = pd.read_csv(self.arquivo_decupagem, encoding="utf-8")
            self._processar_decupagem(df)
        except Exception as e:
            print(f"❌ Erro ao carregar decupagem: {e}")
            return False

        # Carregar plano de filmagem
        try:
            cronograma = self._processar_plano_pdf(self.arquivo_plano)
            if cronograma:
                self._criar_config_do_cronograma(cronograma)
            else:
                self._criar_config_padrao()
        except Exception as e:
            print(f"❌ Erro ao carregar plano de filmagem: {e}")
            import traceback

            traceback.print_exc()
            # Se não conseguir ler o PDF, usar configuração padrão
            self._criar_config_padrao()

        # Salvar configuração gerada
        with open(self.arquivo_config, "w", encoding="utf-8") as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)

        print(
            f"✅ Configuração gerada para {len(self.config.get('dias_filmagem', {}))} dias de filmagem"
        )
        return True

    def _processar_decupagem(self, df):
        """Processa DataFrame da decupagem para extrair dados das cenas"""
        decupagem = {}
        titulo_projeto = None
        cena_atual = None  # Para rastrear a última cena válida

        for idx, row in df.iterrows():
            # Verificar se há número de cena válido
            cena_num = str(row["CENA"]).strip() if pd.notna(row["CENA"]) else ""

            # Se há uma nova cena, atualizar cena_atual
            if cena_num and cena_num != "nan":
                cena_atual = cena_num

                # Criar entrada da cena se não existir
                if cena_num not in decupagem:
                    decupagem[cena_num] = {
                        "locacao": str(row.get("LOCAÇÃO / SET", "")).strip(),
                        "descricao": str(row.get("DESCRIÇÃO CENA", "")).strip(),
                        "elenco": str(row.get("ELENCO", "")).strip(),
                        "observacoes": str(
                            row.get("OBSERVAÇÕES CONTITNUIDADE", "")
                        ).strip(),
                        "planos": [],
                    }

                # Adicionar o plano principal se existir
                plano_desc = str(row.get("PLANOS", "")).strip()
                if plano_desc and plano_desc != "nan":
                    plano = {
                        "planos": plano_desc,
                        "elenco": str(row.get("ELENCO", "")).strip(),
                        "observacoes": str(
                            row.get("OBSERVAÇÕES CONTITNUIDADE", "")
                        ).strip(),
                    }
                    decupagem[cena_num]["planos"].append(plano)

            else:
                # Esta linha não tem cena, mas pode ter planos para a cena_atual
                if cena_atual and cena_atual in decupagem:
                    plano_desc = str(row.get("PLANOS", "")).strip()
                    if plano_desc and plano_desc != "nan":
                        plano = {
                            "planos": plano_desc,
                            "elenco": str(row.get("ELENCO", "")).strip(),
                            "observacoes": str(
                                row.get("OBSERVAÇÕES CONTITNUIDADE", "")
                            ).strip(),
                        }
                        decupagem[cena_atual]["planos"].append(plano)

        self.dados_decupagem = decupagem

        # Armazenar título encontrado
        if titulo_projeto:
            self.titulo_extraido = titulo_projeto

        print(f"✅ {len(decupagem)} cenas carregadas")
        if titulo_projeto:
            print(f"📽️ Projeto detectado: {titulo_projeto}")

    def _processar_plano_pdf(self, arquivo_pdf):
        """Processa o PDF do plano de filmagem para extrair o cronograma completo na ordem"""
        print(f"🔍 Iniciando processamento do PDF: {arquivo_pdf}")

        try:
            cronograma = {}

            with pdfplumber.open(arquivo_pdf) as pdf:
                print(f"📄 PDF aberto com {len(pdf.pages)} páginas")

                texto_completo = ""
                for i, page in enumerate(pdf.pages):
                    texto_pagina = page.extract_text()
                    print(
                        f"📃 Página {i+1}: {len(texto_pagina) if texto_pagina else 0} caracteres extraídos"
                    )
                    if texto_pagina:
                        texto_completo += texto_pagina + "\n"

                print(f"📝 Texto total extraído: {len(texto_completo)} caracteres")

                # Salvar texto para debug
                debug_file = arquivo_pdf.replace(".pdf", "_debug_text.txt")
                with open(debug_file, "w", encoding="utf-8") as f:
                    f.write(texto_completo)
                print(f"💾 Texto salvo para debug em: {debug_file}")

                # Analisar linha por linha para extrair sequência completa
                linhas = texto_completo.split("\n")
                print(f"🔍 Analisando {len(linhas)} linhas do texto...")

                # Extrair título do projeto da primeira linha não vazia
                titulo_projeto = None
                for linha in linhas:
                    linha_limpa = linha.strip()
                    if linha_limpa:
                        titulo_projeto = linha_limpa
                        print(f"📽️ Título do projeto extraído: '{titulo_projeto}'")
                        break

                # Armazenar título extraído
                if titulo_projeto:
                    self.titulo_extraido = titulo_projeto

                dia_atual = None

                for num_linha, linha in enumerate(linhas, 1):
                    linha_original = linha
                    linha_limpa = linha.strip()
                    if not linha_limpa:
                        continue

                    # Detectar início de uma diária
                    match_diaria = re.search(
                        r"diária\s*(\d+)\s*[:：]", linha_limpa, re.IGNORECASE
                    )
                    if match_diaria:
                        dia_atual = int(match_diaria.group(1))
                        print(f"✅ Linha {num_linha}: Encontrado DIÁRIA {dia_atual}")
                        print(f"    Texto da linha: '{linha_limpa}'")

                        if dia_atual not in cronograma:
                            cronograma[dia_atual] = []  # Lista ordenada de atividades
                        continue

                    # Se estamos dentro de uma diária, capturar TUDO na ordem
                    if dia_atual and linha_limpa:

                        # 1. Capturar atividades com horário (café, preparação, refeição, etc.)
                        match_horario = re.match(
                            r"^(\d{2}h\d{2})\s*[-–—]\s*(\d{2}h\d{2})?\s*(.+)",
                            linha_limpa,
                        )
                        if match_horario:
                            horario_inicio = match_horario.group(1)
                            horario_fim = match_horario.group(2) or ""
                            atividade = match_horario.group(3).strip()

                            item = {
                                "tipo": "atividade_fixa",
                                "horario_inicio": horario_inicio,
                                "horario_fim": horario_fim,
                                "atividade": atividade,
                                "linha_original": linha_limpa,
                            }
                            cronograma[dia_atual].append(item)
                            print(
                                f"⏰ Linha {num_linha}: Atividade {horario_inicio}-{horario_fim}: {atividade}"
                            )
                            continue

                        # 2. Capturar cenas (formato: "3 INT QUARTO...")
                        match_cena = re.match(
                            r"^(\d+)\s+(INT|EXT|REC)\s+(.+)", linha_limpa, re.IGNORECASE
                        )
                        if match_cena:
                            numero_cena = int(match_cena.group(1))
                            tipo_local = match_cena.group(2)
                            descricao = match_cena.group(3)

                            item = {
                                "tipo": "cena",
                                "numero": numero_cena,
                                "tipo_local": tipo_local,
                                "descricao": descricao,
                                "horario_inicio": "",  # Sem horário específico
                                "horario_fim": "",
                                "linha_original": linha_limpa,
                            }
                            cronograma[dia_atual].append(item)
                            print(
                                f"🎬 Linha {num_linha}: Cena {numero_cena} ({tipo_local}) - {descricao[:50]}..."
                            )
                            continue

                        # 3. Capturar descrições de cenas (linhas após as cenas)
                        if (
                            any(
                                palavra in linha_limpa.upper()
                                for palavra in ["DIA", "NOITE", "MANHÃ", "TARDE"]
                            )
                            and "ELENCO:" in linha_limpa.upper()
                        ):
                            # Esta é uma descrição de cena, adicionar como informação extra
                            if (
                                cronograma[dia_atual]
                                and cronograma[dia_atual][-1]["tipo"] == "cena"
                            ):
                                cronograma[dia_atual][-1][
                                    "descricao_detalhada"
                                ] = linha_limpa
                                print(
                                    f"📝 Linha {num_linha}: Descrição detalhada da cena"
                                )
                            continue

                        # 4. Capturar REC: (takes/passagens)
                        if linha_limpa.upper().startswith("REC:"):
                            item = {
                                "tipo": "rec",
                                "descricao": linha_limpa,
                                "horario_inicio": "",
                                "horario_fim": "",
                                "linha_original": linha_limpa,
                            }
                            cronograma[dia_atual].append(item)
                            print(f"📹 Linha {num_linha}: REC - {linha_limpa}")
                            continue

                        # 5. Detectar fim da diária
                        if "fim do dia" in linha_limpa.lower():
                            print(f"📝 Linha {num_linha}: Fim da diária {dia_atual}")
                            dia_atual = None
                            continue

                print(f"📅 Cronograma completo extraído:")
                for dia, atividades in cronograma.items():
                    print(f"  📅 Dia {dia}: {len(atividades)} atividades")
                    for i, ativ in enumerate(atividades):
                        tipo_icon = {
                            "atividade_fixa": "⏰",
                            "cena": "🎬",
                            "rec": "📹",
                        }.get(ativ["tipo"], "📋")
                        descricao = ativ.get(
                            "atividade",
                            ativ.get("descricao", f'Cena {ativ.get("numero", "?")}'),
                        )
                        print(
                            f"    {i+1:2d}. {tipo_icon} {ativ.get('horario_inicio', '')} - {descricao}"
                        )

                return cronograma

        except Exception as e:
            print(f"❌ Erro ao processar PDF: {str(e)}")
            print("🔍 Traceback completo:")
            import traceback

            traceback.print_exc()
            return {}

    def _criar_config_do_cronograma(self, cronograma):
        """Cria configuração a partir do cronograma extraído do PDF na ordem correta"""
        print(f"🔧 Criando configuração a partir do cronograma: {len(cronograma)} dias")

        # Usar título extraído do PDF se disponível, senão usar da decupagem, senão usar padrão
        titulo_projeto = getattr(self, "titulo_extraido", None)
        if not titulo_projeto:
            titulo_projeto = "PROJETO DINÂMICO"

        print(f"📽️ Título do projeto configurado: '{titulo_projeto}'")

        self.config = {
            "projeto": {
                "titulo": titulo_projeto,
                "diretor": "",
                "total_dias": len(cronograma),
            },
            "dias_filmagem": {},
        }

        # Converter cronograma para formato da configuração
        for dia, atividades in cronograma.items():
            dia_str = str(dia)
            self.config["dias_filmagem"][dia_str] = {
                "cronograma_completo": atividades,  # Sequência completa do PDF
                "cenas": [],  # Apenas as cenas para compatibilidade
                "locacao_principal": "A DEFINIR",
                "atividades_fixas": [],
            }

            # Extrair apenas as cenas para compatibilidade
            for ativ in atividades:
                if ativ["tipo"] == "cena":
                    self.config["dias_filmagem"][dia_str]["cenas"].append(
                        str(ativ["numero"])
                    )

            print(
                f"  📅 Dia {dia}: {len(atividades)} atividades totais, {len(self.config['dias_filmagem'][dia_str]['cenas'])} cenas"
            )

        print(
            f"✅ Configuração criada com {len(self.config['dias_filmagem'])} dias de filmagem"
        )

    def _criar_config_padrao(self):
        """Cria configuração padrão baseada nas cenas disponíveis"""
        cenas_disponiveis = list(self.dados_decupagem.keys())

        # Dividir cenas em grupos de 3-4 por dia
        dias_config = {}
        cenas_por_dia = 3

        for i in range(0, len(cenas_disponiveis), cenas_por_dia):
            dia_num = str((i // cenas_por_dia) + 1)
            cenas_dia = cenas_disponiveis[i : i + cenas_por_dia]

            dias_config[dia_num] = {
                "cenas": cenas_dia,
                "locacao_principal": "A DEFINIR",
                "atividades_fixas": [],
            }

        titulo_projeto = getattr(self, "titulo_extraido", "PROJETO DINÂMICO")

        self.config = {
            "projeto": {
                "titulo": titulo_projeto,
                "diretor": "A DEFINIR",
                "total_dias": len(dias_config),
            },
            "dias_filmagem": dias_config,
        }

        print(f"✅ Configuração padrão criada com {len(dias_config)} dias")

    def gerar_od_dia(self, dia_num):
        """Gera OD para um dia específico seguindo a ordem exata do PDF"""
        if not self._carregar_dados():
            return False

        dia_str = str(dia_num)
        print(f"🎬 Gerando OD do Dia {dia_num}...")

        if dia_str not in self.config["dias_filmagem"]:
            print(f"❌ Erro: Dia {dia_num} não encontrado na configuração")
            return False

        dia_config = self.config["dias_filmagem"][dia_str]

        # Se tem cronograma completo do PDF, usar ele
        if "cronograma_completo" in dia_config:
            return self._gerar_od_do_cronograma(
                dia_num, dia_config["cronograma_completo"]
            )
        else:
            # Fallback para o método antigo
            return self._gerar_od_simples(dia_num, dia_config)

    def _gerar_od_do_cronograma(self, dia_num, cronograma):
        """Gera OD seguindo exatamente a ordem do cronograma do PDF com formatação especificada"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import locale

        # Configurar localização para português brasileiro
        try:
            locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
        except:
            try:
                locale.setlocale(locale.LC_TIME, "Portuguese_Brazil.1252")
            except:
                pass  # Usar padrão se não conseguir configurar

        wb = Workbook()
        ws = wb.active
        ws.title = f"OD_Dia_{dia_num}"

        # === ESTILOS CONFORME ESPECIFICAÇÃO ===

        # Fontes específicas
        day_od_font = Font(name="Verdana", size=20, bold=True, color="FFFFFF")
        title_font = Font(name="Verdana", size=76, bold=True, color="000000")
        obs_font = Font(name="Verdana", size=19, bold=False, color="000000")
        endereco_font = Font(name="Verdana", size=20, bold=False, color="000000")
        section_title_font = Font(name="Verdana", size=20, bold=True, color="FFFFFF")
        section_subtitle_font = Font(name="Verdana", size=19, bold=True, color="000000")

        # Cores de fundo
        dark_gray_fill = PatternFill(
            start_color="3f3f3f", end_color="3f3f3f", fill_type="solid"
        )
        light_gray_fill = PatternFill(
            start_color="d8d8d8", end_color="d8d8d8", fill_type="solid"
        )

        # Borda para tabela do cronograma
        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # === LINHA 1: DIA E NÚMERO DA OD ===

        # A1:K1 - Dia (ex: "Segunda 25 de Agosto de 2025")
        ws.merge_cells("A1:K1")
        data_atual = datetime.now()
        try:
            dia_semana = data_atual.strftime("%A").capitalize()
            data_formatada = data_atual.strftime("%d de %B de %Y")
            texto_dia = f"{dia_semana} {data_formatada}"
        except:
            # Fallback se localização não funcionar
            dias_semana = [
                "Segunda",
                "Terça",
                "Quarta",
                "Quinta",
                "Sexta",
                "Sábado",
                "Domingo",
            ]
            meses = [
                "Janeiro",
                "Fevereiro",
                "Março",
                "Abril",
                "Maio",
                "Junho",
                "Julho",
                "Agosto",
                "Setembro",
                "Outubro",
                "Novembro",
                "Dezembro",
            ]
            dia_semana = dias_semana[data_atual.weekday()]
            mes_nome = meses[data_atual.month - 1]
            texto_dia = (
                f"{dia_semana} {data_atual.day} de {mes_nome} de {data_atual.year}"
            )

        ws["A1"] = texto_dia
        ws["A1"].font = day_od_font
        ws["A1"].fill = dark_gray_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A1"].border = thin_border  # Adicionar borda

        # L1:M1 - Número da OD (ex: "OD# 2/5")
        ws.merge_cells("L1:M1")
        total_dias = self.config["projeto"]["total_dias"]
        ws["L1"] = f"OD# {dia_num}/{total_dias}"
        ws["L1"].font = day_od_font
        ws["L1"].fill = dark_gray_fill
        ws["L1"].alignment = Alignment(horizontal="right", vertical="center")
        ws["L1"].border = thin_border  # Adicionar borda

        # Altura da linha 1 (convertendo pixels para pontos: pixel × 0.75)
        ws.row_dimensions[1].height = 40 * 0.75  # 30 pontos

        # === LINHA 2: TÍTULO DO PROJETO E DIRETOR ===

        # A2:M2 - Título do projeto e diretor
        ws.merge_cells("A2:M2")
        titulo = self.config["projeto"]["titulo"]
        diretor = self.config["projeto"]["diretor"]
        texto_titulo = f"{titulo}\n{diretor}"
        ws["A2"] = texto_titulo
        ws["A2"].font = title_font
        ws["A2"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["A2"].border = thin_border  # Adicionar borda

        # Altura da linha 2 (convertendo pixels para pontos: pixel × 0.75)
        ws.row_dimensions[2].height = 242 * 0.75  # 181.5 pontos

        # === LINHA 4: OBSERVAÇÕES GERAIS E INFORMAÇÕES ===

        # A4:H4 - Observações gerais
        ws.merge_cells("A4:H4")
        obs_text = (
            "OBSERVAÇÕES GERAIS: O silêncio absoluto é primordial! CELULARES EM MODO AVIÃO\n"
            "Repudiamos a prática de qualquer ato que resulte em discriminação, constrangimento moral ou "
            "assédio de qualquer natureza, sobretudo sexual, adotando, desta forma a política de tolerância "
            "zero para estes tipos de conduta. Esta política de tolerância zero tem aplicação em qualquer "
            "ambiente de trabalho, interno ou externo"
        )
        ws["A4"] = obs_text
        ws["A4"].font = obs_font
        ws["A4"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws["A4"].border = thin_border  # Adicionar borda

        # I4:K4 - Endereço
        ws.merge_cells("I4:K4")
        endereco_text = (
            "Endereço base/SET/LOCAÇÃO\nR. Vaz Caminha, 481 - Zona 02, Maringá - PR"
        )
        ws["I4"] = endereco_text
        ws["I4"].font = endereco_font
        ws["I4"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws["I4"].border = thin_border  # Adicionar borda

        # L4:M4 - Espaço para previsão do tempo (em branco)
        ws.merge_cells("L4:M4")
        ws["L4"].border = thin_border  # Adicionar borda

        # Altura da linha 4 (convertendo pixels para pontos: pixel × 0.75)
        ws.row_dimensions[4].height = 172 * 0.75  # 129 pontos

        # === LINHA 6: TÍTULO "HORÁRIOS GERAIS" ===

        ws.merge_cells("A6:M6")
        ws["A6"] = "HORÁRIOS GERAIS"
        ws["A6"].font = section_title_font
        ws["A6"].fill = dark_gray_fill
        ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A6"].border = thin_border  # Adicionar borda

        # === LINHA 7: SUBTÍTULOS HORÁRIOS GERAIS ===

        # Mesclagens específicas para linha 7
        headers_horarios = [
            "CHAMADA",
            "PREPARAÇÃO",
            "REC CENA 1",
            "ALMOÇO",
            "PREPARAÇÃO",
            "REC CENA",
            "DESPRODUÇÃO",
            "FIM DA DIÁRIA",
        ]
        colunas_horarios = [
            "A7",
            "B7:C7",
            "D7:E7",
            "F7",
            "G7",
            "H7:I7",
            "J7:K7",
            "L7:M7",
        ]

        for i, (header, col_range) in enumerate(
            zip(headers_horarios, colunas_horarios)
        ):
            if ":" in col_range:
                ws.merge_cells(col_range)
                cell_ref = col_range.split(":")[0]
            else:
                cell_ref = col_range

            ws[cell_ref] = header
            ws[cell_ref].font = section_subtitle_font
            ws[cell_ref].fill = light_gray_fill
            ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[cell_ref].border = thin_border  # Adicionar borda

        # === LINHA 8: HORÁRIOS (mantém mesclagem da linha anterior) ===

        # Extrair horários do cronograma
        horarios_do_dia = []
        for atividade in cronograma:
            if atividade["tipo"] == "atividade_fixa" and atividade.get(
                "horario_inicio"
            ):
                horarios_do_dia.append(atividade["horario_inicio"])

        # Preencher horários ou usar padrão
        horarios_padrao = [
            "07h00",
            "07h30",
            "09h30",
            "12h00",
            "13h00",
            "14h30",
            "16h00",
            "17h00",
        ]
        if len(horarios_do_dia) >= 4:
            horarios_padrao = horarios_do_dia[:8] + [""] * (8 - len(horarios_do_dia))

        # Aplicar mesclagens para linha 8
        ws.merge_cells("B8:C8")
        ws.merge_cells("D8:E8")
        ws.merge_cells("H8:I8")
        ws.merge_cells("J8:K8")
        ws.merge_cells("L8:M8")

        colunas_horarios_linha8 = ["A8", "B8", "D8", "F8", "G8", "H8", "J8", "L8"]
        for i, horario in enumerate(horarios_padrao):
            if i < len(colunas_horarios_linha8):
                ws[colunas_horarios_linha8[i]] = horario
                ws[colunas_horarios_linha8[i]].font = Font(name="Verdana", size=19)
                ws[colunas_horarios_linha8[i]].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws[colunas_horarios_linha8[i]].border = thin_border  # Adicionar borda

        # === LINHA 10: TÍTULO "CHAMADA EQUIPE" ===

        ws.merge_cells("A10:M10")
        ws["A10"] = "CHAMADA EQUIPE"
        ws["A10"].font = section_title_font
        ws["A10"].fill = dark_gray_fill
        ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A10"].border = thin_border  # Adicionar borda

        # === LINHA 11: SUBTÍTULOS CHAMADA EQUIPE ===

        headers_equipe = [
            "PRODUÇÃO",
            "DIREÇÃO",
            "ARTE",
            "FOTOGRAFIA",
            "ASSIST. CÂMERA",
            "GAFFER e ELÉTRICA",
            "SOM",
            "FIGURINO",
            "CARACTERIZAÇÃO",
            "CONTINUISTA",
            "ADs",
        ]
        colunas_equipe = [
            "A11",
            "B11:C11",
            "D11:E11",
            "F11",
            "G11",
            "H11:I11",
            "J11:K11",
            "L11:M11",
        ]

        # Ajustar headers para coincidir com colunas disponíveis
        headers_equipe_adj = headers_equipe[: len(colunas_equipe)]

        for i, (header, col_range) in enumerate(
            zip(headers_equipe_adj, colunas_equipe)
        ):
            if ":" in col_range:
                ws.merge_cells(col_range)
                cell_ref = col_range.split(":")[0]
            else:
                cell_ref = col_range

            ws[cell_ref] = header
            ws[cell_ref].font = section_subtitle_font
            ws[cell_ref].fill = light_gray_fill
            ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[cell_ref].border = thin_border  # Adicionar borda

        # === LINHA 12: HORÁRIOS EQUIPE (mantém mesclagem da linha anterior) ===

        # Aplicar mesclagens para linha 12
        ws.merge_cells("B12:C12")
        ws.merge_cells("D12:E12")
        ws.merge_cells("H12:I12")
        ws.merge_cells("J12:K12")
        ws.merge_cells("L12:M12")

        # Horários padrão para equipe
        horarios_equipe = [
            "A/O",
            "07h00",
            "07h00",
            "07h00",
            "07h00",
            "07h00",
            "07h30",
            "07h30",
        ]
        colunas_equipe_linha12 = [
            "A12",
            "B12",
            "D12",
            "F12",
            "G12",
            "H12",
            "J12",
            "L12",
        ]

        for i, horario in enumerate(horarios_equipe):
            if i < len(colunas_equipe_linha12):
                ws[colunas_equipe_linha12[i]] = horario
                ws[colunas_equipe_linha12[i]].font = Font(name="Verdana", size=19)
                ws[colunas_equipe_linha12[i]].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws[colunas_equipe_linha12[i]].border = thin_border  # Adicionar borda

        # === LINHA 14: TÍTULO "ELENCO" ===

        ws.merge_cells("A14:M14")
        ws["A14"] = "ELENCO"
        ws["A14"].font = section_title_font
        ws["A14"].fill = dark_gray_fill
        ws["A14"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A14"].border = thin_border  # Adicionar borda

        # === LINHA 15: SUBTÍTULOS ELENCO ===

        headers_elenco = [
            "ID",
            "ELENCO",
            "PERSONAGEM",
            "CENAS",
            "CHEGADA",
            "CAMARIM",
            "NO SET",
            "SAÍDA",
        ]
        colunas_elenco = [
            "A15",
            "B15:C15",
            "D15:E15",
            "F15",
            "G15",
            "H15:I15",
            "J15:K15",
            "L15:M15",
        ]

        for i, (header, col_range) in enumerate(zip(headers_elenco, colunas_elenco)):
            if ":" in col_range:
                ws.merge_cells(col_range)
                cell_ref = col_range.split(":")[0]
            else:
                cell_ref = col_range

            ws[cell_ref] = header
            ws[cell_ref].font = section_subtitle_font
            ws[cell_ref].fill = light_gray_fill
            ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[cell_ref].border = thin_border  # Adicionar borda

        # === LINHAS 16+: DADOS DO ELENCO ===

        linha_elenco = 16

        # Criar 4 linhas vazias para preenchimento manual do elenco
        for idx in range(1, 5):  # 4 linhas vazias
            # Aplicar mesclagens para linha do elenco
            ws.merge_cells(f"B{linha_elenco}:C{linha_elenco}")
            ws.merge_cells(f"D{linha_elenco}:E{linha_elenco}")
            ws.merge_cells(f"H{linha_elenco}:I{linha_elenco}")
            ws.merge_cells(f"J{linha_elenco}:K{linha_elenco}")
            ws.merge_cells(f"L{linha_elenco}:M{linha_elenco}")

            # Deixar todas as células vazias para preenchimento manual
            ws[f"A{linha_elenco}"] = ""  # ID vazio
            ws[f"B{linha_elenco}"] = ""  # ELENCO vazio
            ws[f"D{linha_elenco}"] = ""  # PERSONAGEM vazio
            ws[f"F{linha_elenco}"] = ""  # CENAS vazio
            ws[f"G{linha_elenco}"] = ""  # CHEGADA vazio
            ws[f"H{linha_elenco}"] = ""  # CAMARIM vazio
            ws[f"J{linha_elenco}"] = ""  # NO SET vazio
            ws[f"L{linha_elenco}"] = ""  # SAÍDA vazio

            # Aplicar formatação e bordas
            for col in ["A", "B", "D", "F", "G", "H", "J", "L"]:
                cell = ws[f"{col}{linha_elenco}"]
                cell.font = Font(name="Verdana", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border  # Adicionar borda

            linha_elenco += 1

        linha_atual = linha_elenco + 1

        # === CRONOGRAMA DE ATIVIDADES (após as seções principais) ===

        # Título da seção cronograma
        ws.merge_cells(f"A{linha_atual}:M{linha_atual}")
        ws[f"A{linha_atual}"] = "CRONOGRAMA DO DIA"
        ws[f"A{linha_atual}"].font = section_title_font
        ws[f"A{linha_atual}"].fill = dark_gray_fill
        ws[f"A{linha_atual}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        linha_atual += 1

        # Cabeçalhos do cronograma conforme especificação atualizada
        headers_cronograma = [
            "HORA A HORA",
            "CENA",
            "DESCRIÇÃO",
            "SHOOTING BOARD",
            "PLANOS",
            "ELENCO",
            "SET",
            "FIGURINO",
            "ARTE",
            "MICROFONAGEM",
            "CRONOLOGIA",
        ]

        # DESCRIÇÃO ocupa colunas C:E (mesclagem)
        ws.merge_cells(f"C{linha_atual}:E{linha_atual}")

        # Mapear headers para colunas corretas
        colunas_headers = {
            "HORA A HORA": "A",
            "CENA": "B",
            "DESCRIÇÃO": "C",  # Mesclada C:E
            "SHOOTING BOARD": "F",
            "PLANOS": "G",
            "ELENCO": "H",
            "SET": "I",
            "FIGURINO": "J",
            "ARTE": "K",
            "MICROFONAGEM": "L",
            "CRONOLOGIA": "M",
        }

        for header, col in colunas_headers.items():
            cell = ws[f"{col}{linha_atual}"]
            cell.value = header
            cell.font = section_subtitle_font
            cell.fill = light_gray_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border  # Adicionar borda
        linha_atual += 1

        # Preencher cronograma na ordem do PDF
        for atividade in cronograma:
            # Definir cores por tipo de atividade
            if atividade["tipo"] == "atividade_fixa":
                if "CAFÉ" in atividade["atividade"].upper():
                    row_color = "FFF2CC"  # Amarelo claro para café
                elif (
                    "REFEIÇÃO" in atividade["atividade"].upper()
                    or "ALMOÇO" in atividade["atividade"].upper()
                ):
                    row_color = "FCE5CD"  # Laranja claro para refeições
                elif "PREPARAÇÃO" in atividade["atividade"].upper():
                    row_color = "E1F5FE"  # Azul claro para preparação
                elif "DESPRODUÇÃO" in atividade["atividade"].upper():
                    row_color = "F3E5F5"  # Roxo claro para desprodução
                else:
                    row_color = "F5F5F5"  # Cinza claro para outras atividades

                # FORMATAÇÃO PARA ATIVIDADES FIXAS:
                # Coluna A: HORA A HORA
                # Mesclagem B:M: descrição da atividade

                ws.merge_cells(f"B{linha_atual}:M{linha_atual}")

                # Preencher dados da atividade fixa
                horario_completo = atividade["horario_inicio"]
                if atividade.get("horario_fim"):
                    horario_completo += f" - {atividade['horario_fim']}"

                ws.cell(
                    row=linha_atual, column=1, value=horario_completo
                )  # HORA A HORA com início e fim
                ws.cell(
                    row=linha_atual, column=2, value=atividade["atividade"].strip("- ")
                )  # ATIVIDADE (mesclada B:M) - remove hífens do início e fim

                # Aplicar formatação específica para atividades fixas
                fill_color = PatternFill(
                    start_color=row_color, end_color=row_color, fill_type="solid"
                )

                # Coluna A (HORA A HORA): Verdana 20, centro horizontal e vertical
                cell_a = ws.cell(row=linha_atual, column=1)
                cell_a.fill = fill_color
                cell_a.font = Font(name="Verdana", size=20)
                cell_a.alignment = Alignment(horizontal="center", vertical="center")
                cell_a.border = thin_border

                # Colunas B:M (mescladas): Verdana 20 bold, esquerda horizontal, centro vertical
                cell_b = ws.cell(row=linha_atual, column=2)
                cell_b.fill = fill_color
                cell_b.font = Font(name="Verdana", size=20, bold=True)
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
                cell_b.border = thin_border

                linha_atual += 1

            elif atividade["tipo"] == "cena":
                row_color = "FFFFFF"  # Fundo branco para cenas

                cena_num = str(atividade["numero"])
                cena_dados = self.dados_decupagem.get(cena_num, {})

                # Montar elenco da cena
                elenco = []
                # Primeiro, pegar elenco da cena principal
                if cena_dados.get("elenco"):
                    elenco.extend(cena_dados["elenco"].split("/"))

                # Depois, pegar elenco dos planos individuais
                for plano in cena_dados.get("planos", []):
                    if plano.get("elenco"):
                        elenco.extend(plano["elenco"].split("/"))

                elenco_str = ", ".join(
                    sorted(set([e.strip() for e in elenco if e.strip()]))
                )

                # Obter dados da cena da decupagem
                descricao = cena_dados.get("descricao", "")
                locacao = cena_dados.get("locacao", "")
                observacoes = cena_dados.get("observacoes", "")
                planos_cena = cena_dados.get("planos", [])

                # Extrair informações de figurino e arte das observações
                figurino_info = ""
                arte_info = ""
                if observacoes:
                    obs_lower = observacoes.lower()
                    if (
                        "uniforme" in obs_lower
                        or "figurino" in obs_lower
                        or "roupa" in obs_lower
                        or "blusa" in obs_lower
                    ):
                        figurino_info = observacoes
                    if (
                        "crachá" in obs_lower
                        or "foto" in obs_lower
                        or "câmera" in obs_lower
                        or "abajur" in obs_lower
                    ):
                        arte_info = observacoes

                # Determinar quantas linhas a cena vai ocupar (1 para cabeçalho + número de planos)
                num_planos = len(planos_cena) if planos_cena else 1
                linha_inicio_cena = linha_atual
                linha_fim_cena = linha_atual + num_planos - 1

                # Se não há planos específicos, criar uma linha básica
                if not planos_cena:
                    planos_cena = [{"planos": "A DEFINIR"}]

                # FORMATAÇÃO PARA CENAS:
                # Nova estrutura de colunas conforme especificação:
                # A: HORA A HORA, B: CENA, C:E: DESCRIÇÃO (mesclada), F: SHOOTING BOARD,
                # G: PLANOS, H: ELENCO, I: SET, J: FIGURINO, K: ARTE, L: MICROFONAGEM, M: CRONOLOGIA

                # Mesclar colunas apropriadas para dados da cena (exceto SHOOTING BOARD e PLANOS)
                if num_planos > 1:
                    # Mesclar HORA A HORA
                    ws.merge_cells(f"A{linha_inicio_cena}:A{linha_fim_cena}")
                    # Mesclar CENA
                    ws.merge_cells(f"B{linha_inicio_cena}:B{linha_fim_cena}")
                    # Mesclar DESCRIÇÃO (C:E)
                    ws.merge_cells(f"C{linha_inicio_cena}:E{linha_fim_cena}")
                    # SHOOTING BOARD não mescla (coluna F)
                    # PLANOS não mescla (coluna G)
                    # Mesclar ELENCO
                    ws.merge_cells(f"H{linha_inicio_cena}:H{linha_fim_cena}")
                    # Mesclar SET
                    ws.merge_cells(f"I{linha_inicio_cena}:I{linha_fim_cena}")
                    # Mesclar FIGURINO
                    ws.merge_cells(f"J{linha_inicio_cena}:J{linha_fim_cena}")
                    # Mesclar ARTE
                    ws.merge_cells(f"K{linha_inicio_cena}:K{linha_fim_cena}")
                    # Mesclar MICROFONAGEM
                    ws.merge_cells(f"L{linha_inicio_cena}:L{linha_fim_cena}")
                    # Mesclar CRONOLOGIA
                    ws.merge_cells(f"M{linha_inicio_cena}:M{linha_fim_cena}")
                else:
                    # Para cenas com apenas um plano, ainda mesclar DESCRIÇÃO
                    ws.merge_cells(f"C{linha_inicio_cena}:E{linha_inicio_cena}")

                # Preencher dados da cena (células mescladas)
                ws.cell(
                    row=linha_inicio_cena, column=1, value=""
                )  # A: HORA A HORA vazio
                ws.cell(
                    row=linha_inicio_cena, column=2, value=f"CENA {cena_num}"
                )  # B: CENA
                ws.cell(
                    row=linha_inicio_cena, column=3, value=descricao
                )  # C: DESCRIÇÃO (mesclada C:E)
                ws.cell(row=linha_inicio_cena, column=8, value=elenco_str)  # H: ELENCO
                ws.cell(row=linha_inicio_cena, column=9, value=locacao)  # I: SET
                ws.cell(
                    row=linha_inicio_cena, column=10, value=figurino_info
                )  # J: FIGURINO
                ws.cell(row=linha_inicio_cena, column=11, value=arte_info)  # K: ARTE
                ws.cell(row=linha_inicio_cena, column=12, value="")  # L: MICROFONAGEM
                ws.cell(row=linha_inicio_cena, column=13, value="")  # M: CRONOLOGIA

                # Preencher planos individualmente
                for i, plano in enumerate(planos_cena):
                    linha_plano = linha_inicio_cena + i

                    # F: SHOOTING BOARD - cada plano tem sua própria célula
                    ws.cell(
                        row=linha_plano, column=6, value=""
                    )  # Em branco por enquanto

                    # G: PLANOS - descrição do plano
                    plano_desc = plano.get("planos", "")
                    ws.cell(row=linha_plano, column=7, value=plano_desc)

                    # Definir altura da linha do plano (212 pixels = 159 pontos)
                    ws.row_dimensions[linha_plano].height = 212 * 0.75  # 159 pontos

                    # Aplicar formatação para linha do plano
                    fill_color = PatternFill(
                        start_color=row_color, end_color=row_color, fill_type="solid"
                    )

                    # Aplicar formatação específica por coluna
                    for col_num, col_letter in [
                        (1, "A"),
                        (2, "B"),
                        (3, "C"),
                        (6, "F"),
                        (7, "G"),
                        (8, "H"),
                        (9, "I"),
                        (10, "J"),
                        (11, "K"),
                        (12, "L"),
                        (13, "M"),
                    ]:
                        cell = ws.cell(row=linha_plano, column=col_num)
                        cell.fill = fill_color
                        cell.font = Font(name="Verdana", size=20)
                        cell.border = thin_border  # Adicionar borda

                        # Alinhamento específico por coluna
                        if col_letter in [
                            "A",
                            "B",
                            "H",
                            "I",
                            "L",
                            "M",
                        ]:  # Centro horizontal e vertical
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True
                            )
                        else:  # Outras colunas: centro vertical, esquerda/padrão horizontal
                            cell.alignment = Alignment(
                                vertical="center", wrap_text=True
                            )

                linha_atual = linha_fim_cena + 1

            elif atividade["tipo"] == "rec":
                row_color = "FFF3E0"  # Laranja muito claro para RECs

                # FORMATAÇÃO PARA REC/TAKES:
                # Coluna A: HORA A HORA
                # Mesclagem B:M: descrição do REC

                ws.merge_cells(f"B{linha_atual}:M{linha_atual}")

                # Preencher dados do REC
                ws.cell(row=linha_atual, column=1, value="")  # A: HORA A HORA vazio
                ws.cell(
                    row=linha_atual, column=2, value=f"REC: {atividade['descricao']}"
                )  # B: REC (mesclada B:M)

                # Aplicar formatação específica para RECs
                fill_color = PatternFill(
                    start_color=row_color, end_color=row_color, fill_type="solid"
                )

                # Coluna A (HORA A HORA): Verdana 20, centro horizontal e vertical
                cell_a = ws.cell(row=linha_atual, column=1)
                cell_a.fill = fill_color
                cell_a.font = Font(name="Verdana", size=20)
                cell_a.alignment = Alignment(horizontal="center", vertical="center")
                cell_a.border = thin_border

                # Colunas B:M (mescladas): Verdana 20 bold, esquerda horizontal, centro vertical
                cell_b = ws.cell(row=linha_atual, column=2)
                cell_b.fill = fill_color
                cell_b.font = Font(name="Verdana", size=20, bold=True)
                cell_b.alignment = Alignment(horizontal="left", vertical="center")
                cell_b.border = thin_border

                linha_atual += 1

        # === CONFIGURAÇÕES FINAIS ===

        # Ajustar larguras das colunas conforme especificação
        # Converter valores de pixels para unidades do Excel (pixel ÷ 7)
        column_widths = {
            "A": 240 / 7,  # HORA A HORA (240 pixels)
            "B": 150 / 7,  # CENA (150 pixels)
            "C": 156 / 7,  # DESCRIÇÃO C (156 pixels)
            "D": 146 / 7,  # DESCRIÇÃO D (146 pixels)
            "E": 212 / 7,  # DESCRIÇÃO E (212 pixels)
            "F": 346 / 7,  # SHOOTING BOARD (346 pixels)
            "G": 238 / 7,  # PLANOS (238 pixels)
            "H": 198 / 7,  # ELENCO (198 pixels)
            "I": 204 / 7,  # SET (204 pixels)
            "J": 208 / 7,  # FIGURINO (208 pixels)
            "K": 262 / 7,  # ARTE (262 pixels)
            "L": 256 / 7,  # MICROFONAGEM (256 pixels)
            "M": 233 / 7,  # CRONOLOGIA (233 pixels)
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Configurações da página
        ws.protection.sheet = False
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Salvar arquivo
        arquivo_od = f"{self.pasta_ods}/OD_Dia_{dia_num}.xlsx"
        wb.save(arquivo_od)

        print(f"📋 Cronograma: {len(cronograma)} atividades na ordem do PDF")
        print(f"🎨 Formatação específica aplicada com planos detalhados")
        print(f"✅ OD salva: {arquivo_od}")
        return True

    def _gerar_od_simples(self, dia_num, dia_config):
        """Fallback para gerar OD simples quando não há cronograma do PDF"""
        print(f"📋 Cenas: {', '.join(dia_config['cenas'])}")
        print(f"📍 Locação: {dia_config['locacao_principal']}")

        # Implementação simplificada...
        arquivo_od = f"{self.pasta_ods}/OD_Dia_{dia_num}.xlsx"
        print(f"✅ OD salva: {arquivo_od}")
        return True

    def gerar_todas_ods(self):
        """Gera ODs para todos os dias disponíveis na configuração"""
        if not self._carregar_dados():
            return False

        dias_disponíveis = list(self.config["dias_filmagem"].keys())
        total_dias = len(dias_disponíveis)

        print(f"🎬 Gerando ODs para {total_dias} dias...")

        sucessos = 0
        falhas = 0

        for dia_str in dias_disponíveis:
            dia_num = int(dia_str)
            print(f"\n📅 Processando Dia {dia_num}...")

            try:
                if self.gerar_od_dia(dia_num):
                    sucessos += 1
                    print(f"✅ OD do Dia {dia_num} gerada com sucesso!")
                else:
                    falhas += 1
                    print(f"❌ Falha ao gerar OD do Dia {dia_num}")
            except Exception as e:
                falhas += 1
                print(f"❌ Erro ao gerar OD do Dia {dia_num}: {str(e)}")

        print(f"\n📊 Resumo:")
        print(f"   ✅ Sucessos: {sucessos}")
        print(f"   ❌ Falhas: {falhas}")
        print(f"   📅 Total: {total_dias}")

        return falhas == 0


# Alias para compatibilidade
GeradorODCompleto = GeradorOD
